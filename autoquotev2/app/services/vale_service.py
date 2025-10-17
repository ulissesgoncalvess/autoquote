# app/services/vale_service.py
def executar_robo_vale(data_coleta):
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from openpyxl import Workbook, load_workbook
    from datetime import datetime
    import time
    import os
    import re

    # --- CONFIGURAÇÃO ---
    USER = "emanuele@sevensuprimentos.com.br"
    PASS = "*Eas251080"

    # --- CONVERTE DATA RECEBIDA DO FRONT ---
    # Exemplo: '161025' → 16/10/25
    if not (data_coleta and len(data_coleta.strip()) == 6 and data_coleta.isdigit()):
        raise ValueError("Data inválida! Use o formato DDMMAA, ex: 190825")

    HOJE_str = f"{data_coleta[:2]}/{data_coleta[2:4]}/{data_coleta[4:]}"
    try:
        HOJE = datetime.strptime(HOJE_str, "%d/%m/%y").date()
    except Exception:
        raise ValueError("Data inválida! Use o formato DDMMAA, ex: 190825")

    print(f"📅 Executando robô para data: {HOJE.strftime('%d/%m/%Y')}")

    # --- CAMINHO DO EXCEL ---
    EXCEL_PATH = os.path.join(os.getcwd(), f"planilha_vale_{HOJE.strftime('%d%m%y')}.xlsx")

    # --- LISTA DE UF ---
    ESTADOS = [
        'AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA', 'MT', 'MS', 'MG',
        'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN', 'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO'
    ]

    # --- PREPARA PLANILHA ---
    if os.path.exists(EXCEL_PATH):
        os.remove(EXCEL_PATH)
    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"
    ws.append(["Numero do evento", "UF(VALE)", "DATA", "DESCRIÇÃO", "QTDE", "UNID. MED", "pagina de descrição"])
    wb.save(EXCEL_PATH)

    # 🔥 A partir daqui o resto do seu código Selenium continua exatamente igual


    # --- INICIA SELENIUM ---
    driver = webdriver.Chrome()
    wait = WebDriverWait(driver, 10)
    driver.get("https://vale.coupahost.com/sessions/supplier_login")

    # login
    wait.until(EC.presence_of_element_located((By.ID, "user_login")))
    driver.find_element(By.ID, "user_login").send_keys(USER)
    driver.find_element(By.ID, "user_password").send_keys(PASS, Keys.RETURN)

    # Clica no elemento de data duas vezes
    try:
        time_filter = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ch_start_time"]')))
        time_filter.click()
        time.sleep(5)
        time_filter = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ch_start_time"]')))
        time_filter.click()
    except:
        pass

    # Robo irá buscar todos os casos que a data inicio = data atual, até a ENCONTRAR ONTEM
    encontrou_ontem = False
    while True:
        time.sleep(5)
        tbody = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="quote_request_table_tag"]')))
        print("tabela encontrada")
        linhas = tbody.find_elements(By.TAG_NAME, "tr")
        print(f"✅ Encontradas {len(linhas)} linhas na tabela.")

        # --- Buscar os números do evento ---
        for linha in linhas:
            try:
                colunas = linha.find_elements(By.TAG_NAME, "td")
                if not colunas or len(colunas) < 7:
                    continue

                # pula a linha se existir o ícone amarelo (flag_yellow) em qualquer lugar da linha
                yellow_flags = linha.find_elements(By.CSS_SELECTOR, "img[src*='flag_yellow']")
                if yellow_flags:
                    print("Pulando linha porque contém flag_yellow")
                    continue

                data_inicio_str = colunas[2].text.strip()
                data_inicio = parse_date_str(data_inicio_str)
                if data_inicio is None:
                    print(f"⚠️ Não foi possível interpretar a data '{data_inicio_str}'. Pulando linha.")
                    continue

                if data_inicio < HOJE:
                    encontrou_ontem = True
                    print(f"❌ Encontrou data anterior a HOJE ({HOJE.strftime('%d/%m/%Y')}): {data_inicio.strftime('%d/%m/%Y')}. Parando a coleta.")
                    break

                if data_inicio != HOJE:
                    print(f"⚠️ Data {data_inicio.strftime('%d/%m/%Y')} não é igual a HOJE ({HOJE.strftime('%d/%m/%Y')}). Ignorando linha.")
                    continue
                numero_evento = colunas[0].find_element(By.TAG_NAME, "a").text.strip()
                data_final = colunas[3].text.strip()
                print(f"Número do evento: {numero_evento} | Data final: {data_final}")

                ws.append([numero_evento, '', data_final, '', '', '', ''])

            except Exception as e:
                print(f"⚠️ Não foi possível extrair dados da linha: {e}")

        if encontrou_ontem:
            break

        try:
            proximo = driver.find_element(By.CLASS_NAME, "next_page")
            print("✅ Botão 'Avançar' encontrado, clicando...")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", proximo)
            time.sleep(1)
            proximo.click()
            print("✅ Botão 'Avançar' clicado.")
            time.sleep(3)
        except Exception as e:
            print(f"Não tem mais páginas ou erro ao clicar no botão 'Avançar': {e}")
            break

    wb.save(EXCEL_PATH)
    print(f"💾 Planilha salva em: {EXCEL_PATH}")

    # --- DETALHA CADA EVENTO ---
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Eventos"]

    for row in ws.iter_rows(min_row=2):
        evento = row[0].value
        if not evento:
            continue
        driver.get(f"https://vale.coupahost.com/quotes/external_responses/{evento}/edit")
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))

        # --- VERIFICA EXISTÊNCIA DA PÁGINA DE DESCRIÇÃO ---
        try:
            botoes1 = driver.find_elements(By.XPATH, '//*[@id="pageContentWrapper"]/div[3]/div[2]/a[2]/span')
            if not botoes1:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                botoes2 = driver.find_elements(By.ID, 'quote_response_submit')
                if botoes2:
                    botoes2[0].click()
        except Exception:
            row[6].value = "Erro ao verificar página de descrição"

        # Scroll e abre seção das informações
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "s-expandLines")))
        elementos = driver.find_elements(By.CLASS_NAME, "s-expandLines")

        if not elementos:
            print(f"⚠️ Nenhum s-expandLines encontrado no evento {evento}")
            continue

        # Duplicar a linha do evento pelo número de elementos encontrados
        linhas_evento = [row]
        if len(elementos) > 1:
            for i in range(len(elementos) - 1):
                nova_linha = [evento, row[1].value, row[2].value, '', '', '', '']
                ws.append(nova_linha)
            wb.save(EXCEL_PATH)
            linhas_evento = [r for r in ws.iter_rows(min_row=2) if r[0].value == evento]

        # Percorre cada s-expandLines e coleta os dados (re-fetch a cada iteração, marca processed via JS)
        def click_element_retry(el, attempts=4, pause=0.4):
            from selenium.common.exceptions import (
                StaleElementReferenceException,
                ElementClickInterceptedException,
                ElementNotInteractableException,
                WebDriverException,
            )
            for _ in range(attempts):
                try:
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    time.sleep(0.15)
                    el.click()
                    return True
                except (StaleElementReferenceException, ElementClickInterceptedException, ElementNotInteractableException, WebDriverException):
                    try:
                        driver.execute_script("arguments[0].click();", el)
                        return True
                    except Exception:
                        time.sleep(pause)
            return False

        # determina quantos existem no DOM no momento (evita usar lista obsoleta)
        total = driver.execute_script("return document.querySelectorAll('.s-expandLines').length")
        if total == 0:
            print(f"⚠️ Nenhum s-expandLines encontrado no evento {evento}")
            continue

        # duplicar linha já feito acima; garante linhas_evento atualizado
        linhas_evento = [r for r in ws.iter_rows(min_row=2) if r[0].value == evento]

        processed = 0
        max_attempts_per_index = 5
        idx = 0
        while processed < total and idx < total:
            # re-obtem a lista sempre
            try:
                elementos = driver.find_elements(By.CLASS_NAME, "s-expandLines")
            except Exception:
                time.sleep(0.3)
                elementos = driver.find_elements(By.CLASS_NAME, "s-expandLines")

            if idx >= len(elementos):
                # DOM encolheu — tenta refetch algumas vezes
                retry_try = 0
                while retry_try < 3 and idx >= len(elementos):
                    time.sleep(0.4)
                    elementos = driver.find_elements(By.CLASS_NAME, "s-expandLines")
                    retry_try += 1
                if idx >= len(elementos):
                    print(f"⚠️ Índice {idx} fora do range atual ({len(elementos)}). Pulando.")
                    idx += 1
                    continue

            el = elementos[idx]

            # evita re-processar elemento já marcado
            already = driver.execute_script("return arguments[0].getAttribute('data-processed')", el)
            if already:
                idx += 1
                processed += 1
                continue

            # tenta clicar de forma robusta
            if not click_element_retry(el, attempts=4, pause=0.4):
                print(f"⚠️ Falha ao clicar no expandLines index {idx} do evento {evento}")
                # marca como processado para não travar loop
                try:
                    driver.execute_script("arguments[0].setAttribute('data-processed','1')", el)
                except Exception:
                    pass
                idx += 1
                processed += 1
                continue

            # após clique, espera conteúdo de detalhe carregar (xpath de descrição)
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]')))
                time.sleep(0.25)
            except Exception:
                time.sleep(0.4)

            # atualiza linhas_evento porque podem ter sido adicionadas
            linhas_evento = [r for r in ws.iter_rows(min_row=2) if r[0].value == evento]
            try:
                linha_atual = linhas_evento[idx]
            except Exception:
                # se não existir, tenta mapear para próxima disponível
                if linhas_evento:
                    linha_atual = linhas_evento[-1]
                else:
                    print(f"⚠️ Não há linha disponível para evento {evento} no idx {idx}")
                    # marca e segue
                    try:
                        driver.execute_script("arguments[0].setAttribute('data-processed','1')", el)
                    except Exception:
                        pass
                    idx += 1
                    processed += 1
                    continue

            # coleta campos (mesma lógica, com pequenos waits)
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[2]/div/div[2]/div/p/span[1]')))
            except Exception:
                time.sleep(1)
            try:
                quantidade_el = driver.find_element(By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[2]/div/div[2]/div/p/span[1]')
                linha_atual[4].value = quantidade_el.text
            except Exception:
                linha_atual[4].value = 'Não foi possivel coletar a quantidade'

            try:
                unidade_el = driver.find_element(By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[2]/div/div[2]/div/p/span[2]')
                linha_atual[5].value = unidade_el.text
            except Exception:
                linha_atual[5].value = 'Não foi possivel coletar a unidade'

            try:
                descri_el = driver.find_element(By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[1]/div/div[2]/div/p')
                descri = descri_el.text
                desejado = re.search(r'PT\s*\|\|\s*(.*?)\*{3,}', descri, re.DOTALL)
                linha_atual[3].value = desejado.group(1).strip() if desejado else descri
            except Exception:
                linha_atual[3].value = 'Não foi possivel coletar a descrição'

            # UF (versão mais robusta)
            try:
                uf_spans = driver.find_elements(
                    By.XPATH,
                    '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[1]/div/div[8]/div/ul/li/span'
                )

                found = None
                # tenta primeiro padrão explícito "- XX - BR" em cada span
                for elem in uf_spans:
                    text = (elem.text or "").strip().upper()
                    if not text:
                        continue
                    m = re.search(r'-\s*([A-Z]{2})\s*-\s*BR', text)
                    if m and m.group(1) in ESTADOS:
                        found = m.group(1)
                        break
                    # procura tokens isolados de 2 letras e valida contra ESTADOS
                    tokens = re.findall(r'\b[A-Z]{2}\b', text)
                    for t in tokens:
                        if t in ESTADOS:
                            found = t
                            break
                    if found:
                        break

                # fallback: junta todo o texto e procura por siglas com bordas de palavra
                if not found:
                    combined = " ".join([(e.text or "") for e in uf_spans]).upper()
                    for sig in ESTADOS:
                        if re.search(r'\b' + re.escape(sig) + r'\b', combined):
                            found = sig
                            break

                linha_atual[1].value = found if found else 'UF não encontrada'
            except Exception:
                linha_atual[1].value = 'Não foi possivel coletar a UF'

            # fecha o detalhe (tenta vários métodos)
            try:
                time.sleep(0.2)
                fechar = None
                try:
                    fechar = driver.find_element(By.CSS_SELECTOR, "button.button.s-cancel")
                except Exception:
                    try:
                        fechar = driver.find_element(By.XPATH, "//button[contains(concat(' ', normalize-space(@class), ' '), ' s-cancel ') and contains(., 'Cancelar')]")
                    except Exception:
                        fechar = None
                if fechar:
                    click_element_retry(fechar, attempts=3, pause=0.2)
                    time.sleep(0.25)
            except Exception:
                pass

            # marca como processado (para não reprocessar se DOM reorganizar)
            try:
                driver.execute_script("arguments[0].setAttribute('data-processed','1')", el)
            except Exception:
                pass

            processed += 1
            idx += 1

        wb.save(EXCEL_PATH)
    # Ordena a planilha por "Numero do evento" (coluna A) para agrupar linhas com o mesmo número
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Eventos"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        def sort_key(row):
            v = row[0]
            if v is None:
                return (1, "")
            s = str(v).strip()
            try:
                return (0, int(s))      # números antes de strings, ordenados numericamente
            except Exception:
                return (1, s.lower())   # strings ordenadas alfabeticamente

        rows_sorted = sorted(rows, key=sort_key)

        # remove linhas antigas (todas a partir da linha 2) e escreve ordenado
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for r in rows_sorted:
            ws.append(list(r))

        wb.save(EXCEL_PATH)
    except Exception as e:
        print(f"⚠️ Falha ao ordenar a planilha: {e}")

    try:
        driver.quit()
    except:
        pass
    print(f"🚀 Iniciando robô para data: {data_coleta}")
    
    # 🔥 AQUI ENTRA SEU CÓDIGO COMPLETO:
    # - Login no portal Vale
    # - Coleta de eventos
    # - Geração do Excel
    # - Tratamento de erros
    
    return {
        "sucesso": True,
        "arquivo": "planilha_vale.xlsx",
        "eventos_coletados": 15,
        "logs": "Robô executado com sucesso!"
    }