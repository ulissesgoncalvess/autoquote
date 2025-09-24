import ttkbootstrap as tb
from ttkbootstrap.constants import *
import subprocess
import threading
import os
from PIL import Image, ImageTk  # Para manipular imagem
from tkinter import BOTH
import sys

def executar_funcao():
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from openpyxl import Workbook, load_workbook
    from datetime import date, timedelta, datetime
    from tkinter.filedialog import asksaveasfilename
    import time
    import os
    import re
    import sys
    import tkinter as tk
    from tkinter import simpledialog

    # --- CONFIGURAÇÃO ---
    EXCEL_PATH = asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Salvar planilha como"
    )
    USER = "emanuele@sevensuprimentos.com.br"
    PASS = "*Eas251080"

    root = tk.Tk()
    root.withdraw()
    data_usuario = simpledialog.askstring(
        title="Input",
        prompt="Digite a data desejada no formato DDMMAA (ex: 190825 para 19/08/25):"
    )
    if not (data_usuario and len(data_usuario.strip()) == 6 and data_usuario.isdigit()):
        raise ValueError("Data inválida! Use o formato DDMMAA, ex: 190825")

    # transforma input DDMMAA em string legível e em objeto date para comparações corretas
    HOJE_str = f"{data_usuario[:2]}/{data_usuario[2:4]}/{data_usuario[4:]}"
    
    def parse_date_str(s: str):
        """Tenta vários formatos e retorna datetime.date ou None."""
        for fmt in ("%d/%m/%y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s.strip(), fmt).date()
            except Exception:
                continue
        return None

    HOJE = parse_date_str(HOJE_str)
    if not HOJE:
        raise ValueError("Data inválida após parse. Use DDMMAA ou DDMMAAAA.")
    root.destroy()

    ESTADOS = [
        'AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA', 'MT', 'MS', 'MG',
        'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN', 'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO'
    ]

    # --- PREPARA PLANILHA ---
    if os.path.exists(EXCEL_PATH):
        os.remove(EXCEL_PATH)
    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"
    ws.append(["Numero do evento", "UF(VALE)", "DATA", "DESCRIÇÃO", "QTDE", "UNID. MED", "pagina de descrição"])
    wb.save(EXCEL_PATH)

    # --- INICIA SELENIUM ---
    driver = webdriver.Chrome()
    wait = WebDriverWait(driver, 10)
    driver.get("https://vale.coupahost.com/sessions/supplier_login")

    # login
    wait.until(EC.presence_of_element_located((By.ID, "user_login")))
    driver.find_element(By.ID, "user_login").send_keys(USER)
    driver.find_element(By.ID, "user_password").send_keys(PASS, Keys.RETURN)

    # Clica no elemento de data duas vezes
    try:
        time_filter = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ch_start_time"]')))
        time_filter.click()
        time.sleep(5)
        time_filter = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ch_start_time"]')))
        time_filter.click()
    except:
        pass

    # Robo irá buscar todos os casos que a data inicio = data atual, até a ENCONTRAR ONTEM
    encontrou_ontem = False
    while True:
        time.sleep(5)
        tbody = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="quote_request_table_tag"]')))
        print("tabela encontrada")
        linhas = tbody.find_elements(By.TAG_NAME, "tr")
        print(f"✅ Encontradas {len(linhas)} linhas na tabela.")

        # --- Buscar os números do evento ---
        for linha in linhas:
            try:
                colunas = linha.find_elements(By.TAG_NAME, "td")
                if not colunas or len(colunas) < 7:
                    continue

                # pula a linha se existir o ícone amarelo (flag_yellow) em qualquer lugar da linha
                yellow_flags = linha.find_elements(By.CSS_SELECTOR, "img[src*='flag_yellow']")
                if yellow_flags:
                    print("Pulando linha porque contém flag_yellow")
                    continue

                data_inicio_str = colunas[2].text.strip()
                data_inicio = parse_date_str(data_inicio_str)
                if data_inicio is None:
                    print(f"⚠️ Não foi possível interpretar a data '{data_inicio_str}'. Pulando linha.")
                    continue

                if data_inicio < HOJE:
                    encontrou_ontem = True
                    print(f"❌ Encontrou data anterior a HOJE ({HOJE.strftime('%d/%m/%Y')}): {data_inicio.strftime('%d/%m/%Y')}. Parando a coleta.")
                    break

                if data_inicio != HOJE:
                    print(f"⚠️ Data {data_inicio.strftime('%d/%m/%Y')} não é igual a HOJE ({HOJE.strftime('%d/%m/%Y')}). Ignorando linha.")
                    continue
                numero_evento = colunas[0].find_element(By.TAG_NAME, "a").text.strip()
                data_final = colunas[3].text.strip()
                print(f"Número do evento: {numero_evento} | Data final: {data_final}")

                ws.append([numero_evento, '', data_final, '', '', '', ''])

            except Exception as e:
                print(f"⚠️ Não foi possível extrair dados da linha: {e}")

        if encontrou_ontem:
            break

        try:
            proximo = driver.find_element(By.CLASS_NAME, "next_page")
            print("✅ Botão 'Avançar' encontrado, clicando...")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", proximo)
            time.sleep(1)
            proximo.click()
            print("✅ Botão 'Avançar' clicado.")
            time.sleep(3)
        except Exception as e:
            print(f"Não tem mais páginas ou erro ao clicar no botão 'Avançar': {e}")
            break

    wb.save(EXCEL_PATH)
    print(f"💾 Planilha salva em: {EXCEL_PATH}")

    # --- DETALHA CADA EVENTO ---
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Eventos"]

    for row in ws.iter_rows(min_row=2):
        evento = row[0].value
        if not evento:
            continue
        driver.get(f"https://vale.coupahost.com/quotes/external_responses/{evento}/edit")
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))

        # --- VERIFICA EXISTÊNCIA DA PÁGINA DE DESCRIÇÃO ---
        try:
            botoes1 = driver.find_elements(By.XPATH, '//*[@id="pageContentWrapper"]/div[3]/div[2]/a[2]/span')
            if not botoes1:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                botoes2 = driver.find_elements(By.ID, 'quote_response_submit')
                if botoes2:
                    botoes2[0].click()
        except Exception:
            row[6].value = "Erro ao verificar página de descrição"

        # Scroll e abre seção das informações
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "s-expandLines")))
        elementos = driver.find_elements(By.CLASS_NAME, "s-expandLines")

        if not elementos:
            print(f"⚠️ Nenhum s-expandLines encontrado no evento {evento}")
            continue

        # Duplicar a linha do evento pelo número de elementos encontrados
        linhas_evento = [row]
        if len(elementos) > 1:
            for i in range(len(elementos) - 1):
                nova_linha = [evento, row[1].value, row[2].value, '', '', '', '']
                ws.append(nova_linha)
            wb.save(EXCEL_PATH)
            linhas_evento = [r for r in ws.iter_rows(min_row=2) if r[0].value == evento]

        # Percorre cada s-expandLines e coleta os dados
        for idx, el in enumerate(elementos):
            try:
                wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "s-expandLines")))
                el.click()
                time.sleep(1)  # pequena pausa para expandir
            except Exception:
                print(f"⚠️ Não consegui clicar no expandLines {idx}")
                continue

            # segura a linha atual (cada linha é uma tuple de células)
            try:
                linha_atual = linhas_evento[idx]
            except IndexError:
                print(f"⚠️ Índice {idx} fora de range para linhas_evento do evento {evento}")
                continue

            # quantidade
            try:
                quantidade = driver.find_element(By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[2]/div/div[2]/div/p/span[1]').text
                linha_atual[4].value = quantidade
            except Exception:
                linha_atual[4].value = 'Não foi possivel coletar a quantidade'

            # unidade
            try:
                unidade = driver.find_element(By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[2]/div/div[2]/div/p/span[2]').text
                linha_atual[5].value = unidade
            except Exception:
                linha_atual[5].value = 'Não foi possivel coletar a unidade'

            # descrição
            try:
                descri = driver.find_element(By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[1]/div/div[2]/div/p').text
                desejado = re.search(r'PT\s*\|\|\s*(.*?)\*{3,}', descri, re.DOTALL)
                if desejado:
                    linha_atual[3].value = desejado.group(1).strip()
                else:
                    linha_atual[3].value = descri
            except Exception:
                linha_atual[3].value = 'Não foi possivel coletar a descrição'

            # UF
            try:
                uf_text = driver.find_element(By.XPATH, '//*[@id="itemsAndServicesApp"]/div/div/div[1]/div[2]/div[2]/div/form/div/div/div[1]/div/div[8]/div/ul/li[1]/span').text
                for sig in ESTADOS:
                    if sig in uf_text:
                        linha_atual[1].value = sig
                        break
            except Exception:
                linha_atual[1].value = 'Não foi possivel coletar a UF'

                #fechar
            try:
              time.sleep(1)
              # Localiza o botão
              fechar = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "button.button.s-cancel")))
              # Rola até o botão estar visível
              driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", fechar)
    
              #  Espera até ele estar clicável
              fechar = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.button.s-cancel")))
              # Clica no botão
              fechar.click()
              time.sleep(1)
            except Exception:
             pass
        wb.save(EXCEL_PATH)

    # Ordena a planilha por "Numero do evento" (coluna A) para agrupar linhas com o mesmo número
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Eventos"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        def sort_key(row):
            v = row[0]
            if v is None:
                return (1, "")
            s = str(v).strip()
            try:
                return (0, int(s))      # números antes de strings, ordenados numericamente
            except Exception:
                return (1, s.lower())   # strings ordenadas alfabeticamente

        rows_sorted = sorted(rows, key=sort_key)

        # remove linhas antigas (todas a partir da linha 2) e escreve ordenado
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for r in rows_sorted:
            ws.append(list(r))

        wb.save(EXCEL_PATH)
    except Exception as e:
        print(f"⚠️ Falha ao ordenar a planilha: {e}")

    try:
        driver.quit()
    except:
        pass


# --- INTERFACE ---
janela = tb.Window(themename="flatly")
janela.title("Robô de Eventos - Seven")
janela.geometry("800x400")
janela.resizable

frame = tb.Frame(janela, padding=20)
frame.pack(fill=BOTH, expand=True)

# Título
titulo = tb.Label(frame, text="Robô de Eventos Seven", font=("Segoe UI", 18, "bold"))
titulo.pack(pady=(0, 20))

# Botão de iniciar
botao_iniciar = tb.Button(frame, text="Iniciar Robô", bootstyle=SUCCESS, width=30, command=executar_funcao)
botao_iniciar.pack(pady=5)

# Status
status_var = tb.StringVar(value="Aguardando início...")
status_label = tb.Label(frame, textvariable=status_var, bootstyle=INFO)
status_label.pack(pady=(20, 0))

janela.mainloop()